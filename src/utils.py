import csv
import json
import logging
from pathlib import Path
from typing import Any, Dict, List, Union

from openpyxl import load_workbook

logger = logging.getLogger('utils')
logger.setLevel(logging.DEBUG)

file_handler = logging.FileHandler('utils.log', encoding='utf-8')
file_formatter = logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)


def _validate_transaction(item: dict) -> bool:
    """Валидация структуры транзакции"""
    required_fields = {
        'date',
        'description',
        'state',
        'operationAmount'
    }

    if not all(field in item for field in required_fields):
        return False

    amount_info = item.get('operationAmount')
    if not isinstance(amount_info, dict):
        return False

    return 'amount' in amount_info and 'currency' in amount_info


def load_json_transactions(file_path: Path) -> List[Dict[str, Any]]:
    """Загрузка данных из JSON-файла"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, list):
            return []
        return [item for item in data if _validate_transaction(item)]
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {str(e)}", exc_info=True)
        return []
    except OSError as e:
        logger.error(f"JSON file error: {str(e)}", exc_info=True)
        return []


def load_csv_transactions(file_path: Path) -> List[Dict[str, Any]]:
    """Загрузка данных из CSV-файла с приведением к общей схеме транзакции."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f, delimiter=';')
            transactions: List[Dict[str, Any]] = []

            for row in reader:
                try:
                    normalized: Dict[str, Any] = {
                        'id': row.get('id'),
                        'date': row.get('date'),
                        'description': row.get('description'),
                        'state': row.get('state'),
                        'from': row.get('from'),
                        'to': row.get('to'),
                        'operationAmount': {
                            'amount': row.get('amount'),
                            'currency': {
                                'name': row.get('currency_name'),
                                'code': row.get('currency_code')
                            }
                        }
                    }

                    if _validate_transaction(normalized):
                        transactions.append(normalized)
                except Exception:
                    continue

            return transactions
    except Exception as e:
        logger.error(f"CSV error: {str(e)}", exc_info=True)
        return []


def load_excel_transactions(file_path: Path) -> List[Dict[str, Any]]:
    """Загрузка данных из Excel-файла"""
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheet = wb.active
        headers = [str(cell.value).strip().lower() for cell in next(sheet.rows)]

        transactions: List[Dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2):
            row_values: Dict[str, Any] = {
                headers[i]: cell.value for i, cell in enumerate(row)
            }

            # Нормализация в общую схему
            normalized: Dict[str, Any] = {
                'id': row_values.get('id'),
                'date': row_values.get('date'),
                'description': row_values.get('description'),
                'state': row_values.get('state'),
                'from': row_values.get('from'),
                'to': row_values.get('to'),
                'operationAmount': {
                    'amount': row_values.get('amount'),
                    'currency': {
                        'name': row_values.get('currency_name'),
                        'code': row_values.get('currency_code')
                    }
                }
            }

            if _validate_transaction(normalized):
                transactions.append(normalized)
        return transactions
    except Exception as e:
        logger.error(f"Excel error: {str(e)}", exc_info=True)
        return []


def load_transactions(file_path: Union[str, Path]) -> List[Dict[str, Any]]:
    """Загружает транзакции с автоматическим определением формата файла"""
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File {path} does not exist")

    if path.suffix.lower() == '.json':
        return load_json_transactions(path)
    elif path.suffix.lower() == '.csv':
        return load_csv_transactions(path)
    elif path.suffix.lower() in ('.xlsx', '.xls'):
        return load_excel_transactions(path)

    raise ValueError(f"Unsupported file format: {path.suffix}")
